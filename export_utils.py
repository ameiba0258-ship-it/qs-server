"""Export utilities for search results."""

import os
import csv
import io
import json
from typing import List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from providers.base import CompanyInfo


def export_to_csv(data: List[CompanyInfo], filepath: str) -> str:
    """Export company data to CSV file. Returns the file path."""
    fieldnames = [
        "company_name", "industry", "mobile_phone", "landline_phone",
        "address", "city", "district", "province",
        "phone", "credit_code", "legal_person", "reg_capital",
        "reg_date", "email", "status"
    ]
    headers_cn = {
        "company_name": "企业名称",
        "industry": "行业",
        "mobile_phone": "手机号",
        "landline_phone": "座机",
        "address": "地址",
        "city": "城市",
        "district": "区县",
        "province": "省份",
        "phone": "联系电话",
        "credit_code": "统一社会信用代码",
        "legal_person": "法定代表人",
        "reg_capital": "注册资本",
        "reg_date": "成立日期",
        "email": "邮箱",
        "status": "经营状态",
    }

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writerow(headers_cn)
        for item in data:
            row = {}
            for k in fieldnames:
                v = getattr(item, k, "")
                row[k] = "; ".join(str(x) for x in v) if isinstance(v, (list, tuple)) else (v or "")
            writer.writerow(row)

    return filepath


def export_to_excel(data: List[CompanyInfo], filepath: str) -> str:
    """Export company data to Excel file with professional formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "企业数据"

    # Headers with Chinese labels
    headers_cn = [
        "企业名称", "行业", "手机号", "座机",
        "地址", "城市", "区县", "省份",
        "联系电话", "统一社会信用代码", "法定代表人", "注册资本",
        "成立日期", "邮箱", "经营状态"
    ]
    fieldnames = [
        "company_name", "industry", "mobile_phone", "landline_phone",
        "address", "city", "district", "province",
        "phone", "credit_code", "legal_person", "reg_capital",
        "reg_date", "email", "status"
    ]

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(size=10, name="微软雅黑")
    data_alignment = Alignment(vertical="center")
    text_alignment = Alignment(vertical="center", horizontal="left")
    thin_border = Border(
        left=Side(style="thin", color="d1d5db"),
        right=Side(style="thin", color="d1d5db"),
        top=Side(style="thin", color="d1d5db"),
        bottom=Side(style="thin", color="d1d5db"),
    )
    alt_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")

    # Write headers
    for col_idx, header in enumerate(headers_cn, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Phone columns (1-indexed): 手机号=3, 座机=4, 联系电话=9
    phone_cols = {3, 4, 9}

    # Data rows
    for row_idx, item in enumerate(data, 2):
        is_alt = row_idx % 2 == 0
        for col_idx, field in enumerate(fieldnames, 1):
            raw = getattr(item, field, "")
            if isinstance(raw, (list, tuple)):
                val = "; ".join(str(v) for v in raw) if raw else ""
            elif raw is None:
                val = ""
            else:
                val = str(raw)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = text_alignment if col_idx > 2 else data_alignment
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill
            # Set phone columns to text format to avoid scientific notation
            if col_idx in phone_cols:
                cell.number_format = '@'

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-adjust column widths with better algorithm
    col_widths = {
        1: 32,   # 企业名称
        2: 18,   # 行业
        3: 16,   # 手机号
        4: 16,   # 座机
        5: 36,   # 地址
        6: 12,   # 城市
        7: 12,   # 区县
        8: 12,   # 省份
        9: 16,   # 联系电话
        10: 22,  # 信用代码
        11: 12,  # 法定代表人
        12: 14,  # 注册资本
        13: 14,  # 成立日期
        14: 22,  # 邮箱
        15: 12,  # 状态
    }
    for col_idx, width in col_widths.items():
        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx - 1) // 26) + chr(65 + (col_idx - 1) % 26)
        ws.column_dimensions[col_letter].width = width

    # Auto-filter for header row
    ws.auto_filter.ref = f"A1:{chr(64 + len(headers_cn))}{len(data) + 1}"

    wb.save(filepath)
    return filepath


def export_json(data: List[CompanyInfo], filepath: str) -> str:
    """Export company data to JSON file."""
    fieldnames = [
        "company_name", "industry", "mobile_phone", "landline_phone",
        "address", "city", "district", "province",
        "phone", "credit_code", "legal_person", "reg_capital",
        "reg_date", "email", "status"
    ]
    rows = [{k: getattr(item, k, "") for k in fieldnames} for item in data]
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2, default=str)
    return filepath
