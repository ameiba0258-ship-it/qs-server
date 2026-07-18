"""POI 商家信息查询工具 - FastAPI Backend"""

import httpx  # noqa: F401

import os
import json
import uuid
import asyncio
import time
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware

from config import settings, get_amap_api_key, get_baidu_api_key, get_tencent_api_key
from providers import get_provider as create_provider
from providers.base import SearchParams, CompanyInfo
from export_utils import export_to_excel, export_to_csv, export_json

import re

def parse_phones(raw_phone: str):
    """Parse phone string into (mobile, landline)."""
    if not raw_phone:
        return "", ""
    parts = re.split(r'[,;，；、/\s]+', str(raw_phone))
    mobiles, landlines = [], []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        cleaned = p.replace('-', '').replace(' ', '')
        if re.match(r'^1[3-9]\d{9}$', cleaned):
            mobiles.append(p)
        elif re.match(r'^0\d{2,3}-?\d{7,8}$', cleaned) or re.match(r'^\d{7,8}$', cleaned):
            landlines.append(p)
        elif re.match(r'^[48]00\d{7,8}$', cleaned):
            landlines.append(p)
        elif re.match(r'^[\d\- ]{7,15}$', p):
            landlines.append(p)
    return "; ".join(mobiles), "; ".join(landlines)

app = FastAPI(title="商家信息查询工具", version="1.0.0")

app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

os.makedirs(settings.EXPORT_DIR, exist_ok=True)

tasks: dict = {}
_providers_cache = {}
BASE_DIR = Path(__file__).parent

_providers_cache = {}

def get_provider_instance(provider_name: str = None):
    global _providers_cache
    if provider_name is None:
        provider_name = "amap"
    
    if provider_name not in _providers_cache:
        key = None
        if provider_name == "amap":
            key = get_amap_api_key()
        elif provider_name == "baidumap":
            key = get_baidu_api_key()
        elif provider_name == "tencentmap":
            key = get_tencent_api_key()
        if not key:
            _providers_cache[provider_name] = None
        else:
            _providers_cache[provider_name] = create_provider(provider_name, key)
    
    return _providers_cache.get(provider_name)

# --- API ---

@app.get("/api/config")
async def api_config():
    return {
        "providers": ["amap", "baidumap", "tencentmap"],
        "current_provider": "amap",
        "api_configured": {
            "amap": bool(get_amap_api_key()),
            "baidumap": bool(get_baidu_api_key()),
            "tencentmap": bool(get_tencent_api_key()),
        },
        "max_results": settings.MAX_RESULTS,
        "page_size": settings.PAGE_SIZE,
    }

@app.post("/api/config/save")
async def save_config(data: dict):
    amap_key = data.get("amap_api_key", "").strip()
    baidu_key = data.get("baidu_api_key", "").strip()
    tencent_key = data.get("tencent_api_key", "").strip()
    if not amap_key and not baidu_key and not tencent_key:
        raise HTTPException(status_code=400, detail="请输入至少一个 API Key")

    # 1) Write to .env file
    env_path = BASE_DIR / ".env"
    try:
        content = env_path.read_text(encoding="utf-8") if env_path.exists() else ""
        lines = content.splitlines(keepends=True)
        found = False
        new_lines = []
        for line in lines:
            if line.startswith("AMAP_API_KEY="):
                new_lines.append(f"AMAP_API_KEY={amap_key}\n")
                found = True
            else:
                new_lines.append(line)
        if not found:
            new_lines.append(f"AMAP_API_KEY={amap_key}\n")
        if amap_key:
            os.environ["AMAP_API_KEY"] = amap_key
        if baidu_key:
            found_b = False
            new_b = []
            for line_in in lines:
                if line_in.startswith("BAIDU_API_KEY="):
                    new_b.append(f"BAIDU_API_KEY={baidu_key}\n")
                    found_b = True
                else:
                    new_b.append(line_in)
            if not found_b:
                new_b.append(f"BAIDU_API_KEY={baidu_key}\n")
            lines = new_b
            os.environ["BAIDU_API_KEY"] = baidu_key
        if tencent_key:
            found_t = False
            new_t = []
            for line_in in lines:
                if line_in.startswith("TENCENT_API_KEY="):
                    new_t.append(f"TENCENT_API_KEY={tencent_key}\n")
                    found_t = True
                else:
                    new_t.append(line_in)
            if not found_t:
                new_t.append(f"TENCENT_API_KEY={tencent_key}\n")
            lines = new_t
            os.environ["TENCENT_API_KEY"] = tencent_key
        env_path.write_text("".join(lines), encoding="utf-8")
        global _providers_caches_cache
        _providers_cache = {}
        return {"success": True, "message": "API Key 已保存并生效"}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
@app.get("/api/regions")
async def get_regions():
    regions_file = BASE_DIR / "data" / "regions.json"
    if regions_file.exists():
        return json.loads(regions_file.read_text(encoding="utf-8"))
    return {"provinces": []}

@app.get("/api/poi-types")
async def get_poi_types():
    poi_file = BASE_DIR / "data" / "poi_types.json"
    if poi_file.exists():
        return json.loads(poi_file.read_text(encoding="utf-8"))
    return {"poi_types": []}

@app.post("/api/search")

async def search_pois(params: SearchParams, request: Request = None):
    # --- Rate limit check ---
    if request:
        token = request.headers.get("Authorization", "").replace("Bearer ", "")
        if token:
            user = auth_db.validate_token(token)
            if user:
                today = auth_db.check_usage(user["id"])
                limit_info = auth_db.TIER_LIMITS.get(user["tier"], auth_db.TIER_LIMITS["free"])
                if today >= limit_info["daily_searches"]:
                    return JSONResponse(status_code=429, content={"success": False, "error": f"今日搜索次数已用完 ({today}/{limit_info['daily_searches']})，升级会员可获更多"})

    provider_name = params.provider or "amap"
    try:
        all_items = []
        seen_set = set()

        if provider_name == "all":
            # Search all configured providers, merge & deduplicate
            providers_to_try = ["amap", "baidumap", "tencentmap"]
            for pn in providers_to_try:
                p = get_provider_instance(pn)
                if not p:
                    continue
                try:
                    r = await p.search(params)
                    for c in r.data:
                        key = c.id or c.company_name
                        if key and key not in seen_set:
                            seen_set.add(key)
                            all_items.append(c)
                        elif not key:
                            all_items.append(c)
                except Exception:
                    continue
            from providers.base import SearchResult
            result = SearchResult(data=all_items, total=len(all_items))
        else:
            provider = get_provider_instance(provider_name)
            if not provider:
                return JSONResponse(status_code=400, content={"success": False, "error": f"数据源 {provider_name} 未配置 API Key"})
            result = await provider.search(params)

        # Parse phone numbers
        data_list = []
        for c in result.data:
            mobile, landline = parse_phones(c.phone)
            data_list.append({
                "id": c.id,
                "company_name": c.company_name,
                "address": c.address,
                "phone": c.phone,
                "mobile_phone": mobile,
                "landline_phone": landline,
                "industry": c.industry,
                "province": c.province,
                "city": c.city,
                "district": c.district,
                "credit_code": c.credit_code,
                "legal_person": c.legal_person,
                "reg_capital": c.reg_capital,
                "reg_date": c.reg_date,
                "email": c.email,
                "status": c.status,
            })

        # Log usage
        username = "anonymous"
        if request:
            token = request.headers.get("Authorization", "").replace("Bearer ", "")
            if token:
                user = auth_db.validate_token(token)
                if user:
                    auth_db.increment_usage(user["id"])
                    username = user["username"]
        
        # Log stats
        try:
            stats_tracker.log_search(params.keyword, params.provider, result.total, username)
        except:
            pass

        return {
            "success": True,
            "total": result.total,
            "page": params.page,
            "page_size": params.page_size,
            "data": data_list,
        }
    except Exception as exc:
        return JSONResponse(status_code=500, content={"success": False, "error": str(exc)})

@app.post("/api/batch-search")
async def batch_search(params: SearchParams, request: Request = None):
    # --- Rate limit check ---
    if request:
        token = request.headers.get("Authorization", "").replace("Bearer ", "")
        if token:
            user = auth_db.validate_token(token)
            if user:
                today = auth_db.check_usage(user["id"])
                limit_info = auth_db.TIER_LIMITS.get(user["tier"], auth_db.TIER_LIMITS["free"])
                if today >= limit_info["daily_searches"]:
                    return JSONResponse(status_code=429, content={"success": False, "error": f"批量搜索次数已用完 ({today}/{limit_info['daily_searches']})"})

    # Check deep_search permission
    deep_search_allowed = True
    _batch_user = None
    if request:
        _token = request.headers.get("Authorization", "").replace("Bearer ", "")
        if _token:
            _u = auth_db.validate_token(_token)
            if _u:
                _batch_user = _u
                _tier_info = auth_db.TIER_LIMITS.get(_u["tier"], auth_db.TIER_LIMITS["free"])
                deep_search_allowed = _tier_info.get("deep_search", False)

    task_id = uuid.uuid4().hex[:12]
    tasks[task_id] = {
        "status": "running", "progress": 0, "total": 0,
        "count": 0, "started_at": time.time(), "error": None, "filepath": None,
    }
    asyncio.create_task(_run_batch_search(task_id, params, deep_search_allowed))
    return {"task_id": task_id}

def _merge_dedup(results_lists, seen_ids=None):
    """Merge multiple result lists, dedup by id."""
    if seen_ids is None:
        seen_ids = set()
    merged = []
    for rlist in results_lists:
        if not rlist:
            continue
        for item in rlist:
            key = item.id or item.company_name
            if key and key not in seen_ids:
                seen_ids.add(key)
                merged.append(item)
            elif not key:
                merged.append(item)
    return merged

async def _deep_search_places(provider, params, places, task_id, max_deep, label):
    """Search each city/district separately and merge deduplicated."""
    all_map = {}
    seen_ids = set()
    for idx, place in enumerate(places):
        if len(all_map) >= max_deep:
            break
        tasks[task_id].update(
            progress_text=f"正在搜索 {place} ({idx+1}/{len(places)}) 已获取 {len(all_map)} 条"
        )
        place_params = SearchParams(
            provider=params.provider, keyword=params.keyword,
            province=params.province, city=place,
            industry=params.industry,
            reg_capital_min=params.reg_capital_min,
            reg_capital_max=params.reg_capital_max,
            reg_date_start=params.reg_date_start,
            reg_date_end=params.reg_date_end,
            page=1, page_size=25
        )
        try:
            place_result = await provider.batch_search(place_params, max_results=100)
        except Exception:
            continue
        for item in place_result:
            if item.id and item.id in seen_ids:
                continue
            if item.id:
                seen_ids.add(item.id)
            all_map[item.id or str(len(all_map))] = item
        tasks[task_id].update(progress=len(all_map), count=len(all_map))
    result = list(all_map.values())
    tasks[task_id].update(progress_text=f"深度搜索完成，共获取 {len(result)} 条（已去重）")
    return result

def _load_regions():
    rpath = BASE_DIR / "data" / "regions.json"
    if rpath.exists():
        return json.loads(rpath.read_text(encoding="utf-8")).get("provinces", [])
    return []

async def _run_single_keyword_search(provider, params, task_id, provinces, MUNICIPALITIES, MAX_DEEP, keyword_label="", deep_search_allowed=True):
    """Run a single keyword search (could be deep or regular). Returns data list."""
    target_city = params.city or params.province
    all_data = None

    can_deep = deep_search_allowed

    if target_city in MUNICIPALITIES and can_deep:
        districts = []
        for p in provinces:
            if p["name"] == target_city:
                districts = p.get("cities", [])
                break
        if districts:
            all_data = await _deep_search_places(provider, params, districts, task_id, MAX_DEEP, "区县")
        else:
            reg_cb = lambda prog, total: tasks[task_id].update(progress=prog, total=total, count=prog)
            all_data = await provider.batch_search(params, max_results=settings.MAX_RESULTS, progress_callback=reg_cb)

    elif params.province and not params.city and can_deep:
        cities = []
        for p in provinces:
            if p["name"] == params.province:
                cities = p.get("cities", [])
                break
        if cities:
            all_data = await _deep_search_places(provider, params, cities, task_id, MAX_DEEP, "城市")
        else:
            reg_cb = lambda prog, total: tasks[task_id].update(progress=prog, total=total, count=prog)
            all_data = await provider.batch_search(params, max_results=settings.MAX_RESULTS, progress_callback=reg_cb)

    else:
        reg_cb = lambda prog, total: tasks[task_id].update(progress=prog, total=total, count=prog)
        all_data = await provider.batch_search(params, max_results=settings.MAX_RESULTS, progress_callback=reg_cb)

    return all_data or []

async def _run_batch_search(task_id: str, params: SearchParams, deep_search_allowed: bool = True):
    provider = get_provider_instance(params.provider or "amap")
    MUNICIPALITIES = {"北京市", "天津市", "上海市", "重庆市"}
    MAX_DEEP = 3000

    try:
        provinces = _load_regions()
        all_keywords = [params.keyword] + params.keywords
        all_keywords = [k.strip() for k in all_keywords if k.strip()]

        if len(all_keywords) > 1:
            # Multi-keyword mode: search each keyword, merge all results
            all_seen = set()
            merged_all = []

            for ki, kw in enumerate(all_keywords):
                if len(merged_all) >= MAX_DEEP:
                    break

                tasks[task_id].update(
                    progress_text=f"关键词 ({ki+1}/{len(all_keywords)}): 「{kw}」 — 当前共 {len(merged_all)} 条"
                )

                kw_params = SearchParams(
                    provider=params.provider, keyword=kw,
                    province=params.province, city=params.city,
                    industry=params.industry,
                    reg_capital_min=params.reg_capital_min,
                    reg_capital_max=params.reg_capital_max,
                    reg_date_start=params.reg_date_start,
                    reg_date_end=params.reg_date_end,
                    page=1, page_size=params.page_size
                )

                kw_results = await _run_single_keyword_search(provider, kw_params, task_id, provinces, MUNICIPALITIES, MAX_DEEP, deep_search_allowed=deep_search_allowed)

                for item in kw_results:
                    key = item.id or item.company_name
                    if key and key not in all_seen:
                        all_seen.add(key)
                        merged_all.append(item)
                    elif not key:
                        merged_all.append(item)

                tasks[task_id].update(progress=len(merged_all), count=len(merged_all))

            all_data = merged_all
            tasks[task_id].update(
                progress_text=f"多关键词搜索完成（{len(all_keywords)}个关键词），共获取 {len(all_data)} 条（已去重）"
            )
        else:
            # Single keyword mode
            all_data = await _run_single_keyword_search(provider, params, task_id, provinces, MUNICIPALITIES, MAX_DEEP, deep_search_allowed=deep_search_allowed)

        # --- Export ---
        if not all_data:
            tasks[task_id].update(status="error", error="没有搜索到数据")
            return

        kw_label = "多关键词" if len(all_keywords) > 1 else ""
        filename = f"{kw_label}POI数据_{time.strftime('%Y%m%d_%H%M%S')}_{task_id}.xlsx"
        filepath = os.path.join(settings.EXPORT_DIR, filename)
        export_to_excel(all_data, filepath)
        tasks[task_id].update(
            status="completed", progress=len(all_data), count=len(all_data),
            filepath=filepath, filename=filename
        )
    except Exception as exc:
        import traceback
        tasks[task_id].update(status="error", error=str(exc))

@app.get("/api/batch-progress/{task_id}")
async def batch_progress(task_id: str):
    task = tasks.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task

@app.get("/api/download/{task_id}")
async def download_result(task_id: str, fmt: str = "xlsx"):
    task = tasks.get(task_id)
    if not task or task["status"] != "completed":
        raise HTTPException(status_code=404, detail="No completed result")
    filepath = task.get("filepath")
    if not filepath or not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="File not found")

    filename = task.get("filename", "result.xlsx")

    if fmt == "csv":
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb.active
        fieldnames = ["company_name", "industry", "mobile_phone", "landline_phone",
                   "address", "city", "district", "province",
                   "phone", "credit_code", "legal_person", "reg_capital",
                   "reg_date", "email", "status"]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v for v in row):
                item = CompanyInfo(**dict(zip(fieldnames, row)))
                data.append(item)
        csv_path = filepath.replace(".xlsx", ".csv")
        export_to_csv(data, csv_path)
        return FileResponse(csv_path, filename=filename.replace(".xlsx", ".csv"), media_type="text/csv; charset=utf-8")
    elif fmt == "json":
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb.active
        fieldnames = ["company_name", "industry", "mobile_phone", "landline_phone",
                   "address", "city", "district", "province",
                   "phone", "credit_code", "legal_person", "reg_capital",
                   "reg_date", "email", "status"]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v for v in row):
                item = CompanyInfo(**dict(zip(fieldnames, row)))
                data.append(item)
        json_path = filepath.replace(".xlsx", ".json")
        export_json(data, json_path)
        return FileResponse(json_path, filename=filename.replace(".xlsx", ".json"), media_type="application/json")
    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/")
async def serve_index():
    """Root now serves the login page. Authenticated users redirect to /search."""
    login_path = BASE_DIR / "static" / "auth" / "login.html"
    if login_path.exists():
        return HTMLResponse(login_path.read_text(encoding="utf-8"))
    return {"message": "Login page not found"}

@app.get("/css/{filename}")
async def serve_css(filename: str):
    filepath = BASE_DIR / "static" / "css" / filename
    if filepath.exists():
        headers = {"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache", "Expires": "0"}
        return FileResponse(str(filepath), media_type="text/css", headers=headers)
    raise HTTPException(status_code=404)

@app.get("/js/{filename}")
async def serve_js(filename: str):
    filepath = BASE_DIR / "static" / "js" / filename
    if filepath.exists():
        headers = {"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache", "Expires": "0"}
        return FileResponse(str(filepath), media_type="application/javascript", headers=headers)
    raise HTTPException(status_code=404)

# --- Auth & Login Routes ---



@app.get("/profile")
async def profile_page():
    """User profile and settings page."""
    profile_path = BASE_DIR / "static" / "profile.html"
    if profile_path.exists():
        return HTMLResponse(profile_path.read_text(encoding="utf-8"))
    return {"message": "Profile page not found"}

@app.get("/search")
async def serve_search():
    """Main search interface — protected by frontend auth check."""
    search_path = BASE_DIR / "static" / "index.html"
    if search_path.exists():
        return HTMLResponse(search_path.read_text(encoding="utf-8"))
    return {"message": "Search page not found"}

@app.get("/login")
async def login_page():
    """Keep /login for backward compatibility — render same page."""
    login_path = BASE_DIR / "static" / "auth" / "login.html"
    if login_path.exists():
        return HTMLResponse(login_path.read_text(encoding="utf-8"))
    return {"message": "Login page not found"}

@app.post("/api/auth/register")
async def register(data: dict):
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    if not username or not password:
        return {"success": False, "error": "用户名和密码不能为空"}
    import auth_db
    result = auth_db.register_user(username, password)
    return result

@app.post("/api/auth/login")
async def login(data: dict):
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    if not username or not password:
        return {"success": False, "error": "用户名和密码不能为空"}
    import auth_db  # noqa: F811
    result = auth_db.login_user(username, password)
    return result

@app.get("/api/auth/profile")
async def profile(request: Request):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return {"success": False, "error": "未登录"}
    import auth_db  # noqa: F811
    user = auth_db.validate_token(token)
    if not user:
        return {"success": False, "error": "登录已过期"}
    info = auth_db.get_user_info(user["id"])
    return {"success": True, "user": info}



@app.post("/api/auth/email-register")
async def email_register(data: dict):
    """Register with email and password."""
    email = data.get("email", "").strip().lower()
    password = data.get("password", "").strip()
    if not email or not password:
        return {"success": False, "error": "邮箱和密码不能为空"}
    if "@" not in email or "." not in email:
        return {"success": False, "error": "请输入有效的邮箱地址"}
    if len(password) < 6:
        return {"success": False, "error": "密码至少6位"}
    import auth_db
    result = auth_db.register_by_email(email, password)
    if result.get("success"):
        result["email"] = email
    return result


@app.post("/api/auth/email-login")
async def email_login(data: dict):
    """Login with email and password."""
    email = data.get("email", "").strip().lower()
    password = data.get("password", "").strip()
    if not email or not password:
        return {"success": False, "error": "邮箱和密码不能为空"}
    import auth_db
    result = auth_db.login_by_email(email, password)
    return result




@app.post("/api/auth/change-password")
async def change_password(data: dict, request: Request):
    """Authenticated user changes their password."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return {"success": False, "error": "未登录"}
    old_password = data.get("old_password", "")
    new_password = data.get("new_password", "")
    if not old_password or not new_password:
        return {"success": False, "error": "请填写原密码和新密码"}
    if len(new_password) < 6:
        return {"success": False, "error": "新密码至少6位"}
    import auth_db
    user = auth_db.validate_token(token)
    if not user:
        return {"success": False, "error": "登录已过期"}
    result = auth_db.change_password(user["id"], old_password, new_password)
    if result.get("success"):
        result["message"] = "密码修改成功，请重新登录"
    return result


@app.post("/api/auth/admin/reset-password")
async def admin_reset_password(data: dict, request: Request):
    """Admin resets a user's password."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return {"success": False, "error": "未登录"}
    import auth_db
    admin_user = auth_db.validate_token(token)
    if not admin_user or admin_user["tier"] != "admin":
        return {"success": False, "error": "无权限"}
    username = data.get("username", "").strip()
    new_password = data.get("new_password", "")
    if not username or not new_password:
        return {"success": False, "error": "请填写用户名和新密码"}
    if len(new_password) < 6:
        return {"success": False, "error": "新密码至少6位"}
    result = auth_db.admin_reset_password(username, new_password)
    return result




@app.get("/api/auth/payment-history")
async def payment_history(request: Request):
    """Get payment history for authenticated user."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return {"success": False, "error": "未登录"}
    import auth_db
    user = auth_db.validate_token(token)
    if not user:
        return {"success": False, "error": "登录已过期"}
    payments = auth_db.get_user_payments(user["username"])
    # Also get pending notifications
    pending_notifs = []
    import json, os
    notif_path = os.path.join(os.path.dirname(__file__), "data", "payments.json")
    if os.path.exists(notif_path):
        with open(notif_path, "r") as f:
            all_notifs = json.load(f)
        pending_notifs = [n for n in all_notifs if n.get("username") == user["username"] and n.get("status") == "pending"]
    return {"success": True, "payments": payments, "pending_notifications": pending_notifs}




@app.get("/api/admin/email-config")
async def get_email_config(request: Request):
    """Admin: get SMTP email config."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return {"success": False, "error": "未登录"}
    import auth_db
    user = auth_db.validate_token(token)
    if not user or user["tier"] != "admin":
        return {"success": False, "error": "无权限"}
    config = auth_db.get_email_config()
    # Don't expose password in full
    safe = dict(config)
    if safe.get("password"):
        safe["password"] = "••••••" if safe["password"] else ""
    return {"success": True, "config": safe}


@app.post("/api/admin/email-config")
async def save_email_config(data: dict, request: Request):
    """Admin: save SMTP email config."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return {"success": False, "error": "未登录"}
    import auth_db
    user = auth_db.validate_token(token)
    if not user or user["tier"] != "admin":
        return {"success": False, "error": "无权限"}
    config = {
        "host": data.get("host", "").strip(),
        "port": int(data.get("port", 587)),
        "username": data.get("username", "").strip(),
        "password": data.get("password", "").strip(),
        "from_addr": data.get("from_addr", "").strip() or data.get("username", "").strip(),
        "configured": bool(data.get("host", "") and data.get("username", "") and data.get("password", "")),
    }
    auth_db.save_email_config(config)
    return {"success": True, "message": "邮件配置已保存"}


# --- Membership & Admin Routes ---

@app.get("/membership")
async def membership_page():
    path = BASE_DIR / "static" / "membership.html"
    if path.exists():
        return HTMLResponse(path.read_text(encoding="utf-8"))
    return {"message": "Page not found"}

@app.get("/api-docs")
async def api_docs_page():
    path = BASE_DIR / "static" / "api" / "docs.html"
    if path.exists():
        return HTMLResponse(path.read_text(encoding="utf-8"))
    return {"message": "Page not found"}

@app.get("/admin")
async def admin_page():
    path = BASE_DIR / "static" / "admin.html"
    if path.exists():
        return HTMLResponse(path.read_text(encoding="utf-8"))
    return {"message": "Page not found"}

@app.post("/api/admin/set-tier")
async def set_tier(data: dict, request: Request):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return JSONResponse(status_code=401, content={"success": False, "error": "未登录"})
    user = auth_db.validate_token(token)
    if not user or user["tier"] != "admin":
        return JSONResponse(status_code=403, content={"success": False, "error": "无权限"})
    
    target_user = data.get("username", "").strip()
    new_tier = data.get("tier", "").strip()
    if not target_user or not new_tier:
        return {"success": False, "error": "参数不完整"}
    
    result = auth_db.set_user_tier(target_user, new_tier)
    return result

@app.get("/api/admin/users")
async def get_users(request: Request):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return JSONResponse(status_code=401, content={"success": False, "error": "未登录"})
    user = auth_db.validate_token(token)
    if not user or user["tier"] != "admin":
        return JSONResponse(status_code=403, content={"success": False, "error": "无权限"})
    users = auth_db.get_all_users()
    return {"success": True, "users": users}

# --- Payment Routes ---

@app.get("/api/payment/config")
async def get_payment_config():
    """Public: get payment config (no auth needed)."""
    import auth_db

    config = auth_db.get_payment_config()
    return {"success": True, "config": config}

@app.post("/api/admin/payment-config")
async def set_payment_config(data: dict, request: Request):
    """Admin: set payment config."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return JSONResponse(status_code=401, content={"success": False, "error": "未登录"})
    import auth_db

    user = auth_db.validate_token(token)
    if not user or user["tier"] != "admin":
        return JSONResponse(status_code=403, content={"success": False, "error": "无权限"})
    
    config = {
        "wechat_id": data.get("wechat_id", ""),
        "alipay_account": data.get("alipay_account", ""),
        "qr_code_url": data.get("qr_code_url", ""),
        "price": data.get("price", "99"),
        "contact": data.get("contact", ""),
    }
    auth_db.save_payment_config(config)
    return {"success": True, "message": "支付设置已保存"}

@app.post("/api/payment/notify")
async def payment_notify(data: dict, request: Request):
    """User submits payment notification."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return JSONResponse(status_code=401, content={"success": False, "error": "请先登录"})
    import auth_db

    user = auth_db.validate_token(token)
    if not user:
        return JSONResponse(status_code=401, content={"success": False, "error": "登录已过期"})
    
    contact = data.get("contact", "")
    amount = data.get("amount", "99")
    notif = auth_db.add_payment_notification(user["username"], amount, contact)
    return {"success": True, "message": "已通知管理员，请等待开通", "notification": notif}

@app.get("/api/admin/payments")
async def get_pending_payments(request: Request):
    """Admin: get pending payments."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return JSONResponse(status_code=401, content={"success": False, "error": "未登录"})
    import auth_db

    user = auth_db.validate_token(token)
    if not user or user["tier"] != "admin":
        return JSONResponse(status_code=403, content={"success": False, "error": "无权限"})
    payments = auth_db.get_pending_payments()
    return {"success": True, "payments": payments}

@app.post("/api/admin/payments/confirm")
async def confirm_user_payment(data: dict, request: Request):
    """Admin confirms payment and upgrades user."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return JSONResponse(status_code=401, content={"success": False, "error": "未登录"})
    import auth_db

    user = auth_db.validate_token(token)
    if not user or user["tier"] != "admin":
        return JSONResponse(status_code=403, content={"success": False, "error": "无权限"})
    
    notif_id = data.get("id")
    target_user = data.get("username", "")
    new_tier = data.get("tier", "premium")
    
    if not notif_id and not target_user:
        return {"success": False, "error": "参数不完整"}
    
    # Confirm payment notification
    if notif_id:
        auth_db.confirm_payment(notif_id)
    
    # Upgrade user
    if target_user:
        result = auth_db.set_user_tier(target_user, new_tier)
        return {"success": True, "message": f"✅ {target_user} 已升级为 {new_tier}", "upgrade": result}
    
    return {"success": True, "message": "付款已确认"}

# --- API Token Management ---

@app.post("/api/admin/api-tokens")
async def create_api_token(data: dict, request: Request):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return JSONResponse(status_code=401, content={"success": False, "error": "未登录"})
    import auth_db

    user = auth_db.validate_token(token)
    if not user or user["tier"] != "admin":
        return JSONResponse(status_code=403, content={"success": False, "error": "无权限"})
    
    tier = data.get("tier", "free")
    new_token = auth_db.create_api_token(user["username"], tier)
    return {"success": True, "token": new_token, "message": f"API Token 已创建（{tier}）"}

@app.get("/api/admin/api-tokens")
async def list_api_tokens(request: Request):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return JSONResponse(status_code=401, content={"success": False, "error": "未登录"})
    import auth_db

    user = auth_db.validate_token(token)
    if not user or user["tier"] != "admin":
        return JSONResponse(status_code=403, content={"success": False, "error": "无权限"})
    tokens = auth_db.get_api_tokens()
    return {"success": True, "tokens": tokens}

@app.delete("/api/admin/api-tokens/{token_id}")
async def delete_api_token(token_id: int, request: Request):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return JSONResponse(status_code=401, content={"success": False, "error": "未登录"})
    import auth_db

    user = auth_db.validate_token(token)
    if not user or user["tier"] != "admin":
        return JSONResponse(status_code=403, content={"success": False, "error": "无权限"})
    auth_db.revoke_api_token(token_id)
    return {"success": True, "message": "API Token 已撤销"}

# --- Stats & Dashboard ---

@app.get("/api/admin/stats")
async def admin_stats(request: Request):
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return JSONResponse(status_code=401, content={"success": False, "error": "未登录"})
    user = auth_db.validate_token(token)
    if not user or user["tier"] != "admin":
        return JSONResponse(status_code=403, content={"success": False, "error": "无权限"})
    
    days = int(request.query_params.get("days", "7"))
    stats = stats_tracker.get_stats(days)
    
    # Add user counts
    users = auth_db.get_all_users()
    tokens = auth_db.get_api_tokens()
    
    return {
        "success": True,
        "stats": stats,
        "total_users": len(users),
        "active_tokens": len([t for t in tokens if t.get("is_active")]),
        "pending_payments": len(auth_db.get_pending_payments()),
    }



# --- Nearby Search ---

@app.get("/api/geocode")
async def geocode_address(address: str, city: str = ""):
    """Geocode an address to coordinates using 高德/腾讯 API."""
    api_key = get_amap_api_key() or get_tencent_api_key()
    if not api_key:
        return {"success": False, "error": "请先配置高德或腾讯 API Key"}
    
    # Try 高德 first
    if get_amap_api_key():
        try:
            async with httpx.AsyncClient(timeout=10) as client:
                resp = await client.get("https://restapi.amap.com/v3/geocode/geo", params={
                    "key": get_amap_api_key(), "address": address, "city": city
                })
                data = resp.json()
            if data.get("status") == "1" and data.get("geocodes"):
                loc = data["geocodes"][0]["location"]
                addr = data["geocodes"][0].get("formatted_address", address)
                return {"success": True, "location": loc, "address": addr, "source": "amap"}
        except:
            pass
    
    # Fallback: 腾讯
    if get_tencent_api_key():
        try:
            async with httpx.AsyncClient(timeout=10) as client:
                resp = await client.get("https://apis.map.qq.com/ws/geocoder/v1/", params={
                    "key": get_tencent_api_key(), "address": address
                })
                data = resp.json()
            if data.get("status") == 0:
                loc = f'{data["result"]["location"]["lng"]},{data["result"]["location"]["lat"]}'
                return {"success": True, "location": loc, "address": address, "source": "tencent"}
        except:
            pass
    
    return {"success": False, "error": "地址解析失败，请检查输入或配置 API Key"}


@app.post("/api/nearby-search")
async def around_search(data: dict):
    """Search nearby a location."""
    location = data.get("location", "")
    keyword = data.get("keyword", "")
    radius = int(data.get("radius", 1000))
    provider_name = data.get("provider", "amap")
    page = int(data.get("page", 1))
    page_size = min(int(data.get("page_size", 20)), 25)
    
    if not location or not keyword:
        return {"success": False, "error": "位置和关键词不能为空"}
    
    results = []
    total = 0
    
    if provider_name == "amap":
        key = get_amap_api_key()
        if not key:
            return {"success": False, "error": "高德 API Key 未配置"}
        try:
            async with httpx.AsyncClient(timeout=15) as client:
                resp = await client.get("https://restapi.amap.com/v3/place/around", params={
                    "key": key, "location": location, "keywords": keyword,
                    "radius": radius, "offset": page_size, "page": page, "extensions": "all"
                })
                d = resp.json()
            if d.get("status") != "1":
                return {"success": False, "error": d.get("info", "搜索失败")}
            total = int(d.get("count", 0))
            for poi in d.get("pois", []):
                mobile, landline = parse_phones(poi.get("tel", ""))
                results.append({
                    "company_name": poi.get("name", ""),
                    "address": poi.get("address", ""),
                    "phone": poi.get("tel", ""),
                    "mobile_phone": mobile,
                    "landline_phone": landline,
                    "industry": poi.get("type", ""),
                    "province": poi.get("pname", ""),
                    "city": poi.get("cityname", ""),
                    "district": poi.get("adname", ""),
                    "distance": poi.get("distance", ""),
                })
        except Exception as e:
            return {"success": False, "error": f"请求失败: {str(e)[:50]}"}
    
    elif provider_name == "baidumap":
        key = get_baidu_api_key()
        if not key:
            return {"success": False, "error": "百度 API Key 未配置"}
        loc_parts = location.split(",")
        if len(loc_parts) == 2:
            lat, lng = loc_parts[1], loc_parts[0]  # Baidu uses lat,lng order
            try:
                async with httpx.AsyncClient(timeout=15) as client:
                    resp = await client.get("https://api.map.baidu.com/place/v2/search", params={
                        "query": keyword, "location": f"{lat},{lng}", "radius": radius,
                        "output": "json", "ak": key, "page_size": page_size, "page_num": page - 1,
                        "scope": "2"
                    })
                    d = resp.json()
                if d.get("status") != 0:
                    return {"success": False, "error": "百度搜索失败"}
                total = int(d.get("total", 0))
                for item in d.get("results", []):
                    tel = item.get("telephone", "")
                    mobile, landline = parse_phones(tel)
                    results.append({
                        "company_name": item.get("name", ""),
                        "address": item.get("address", ""),
                        "phone": tel, "mobile_phone": mobile, "landline_phone": landline,
                        "industry": item.get("detail", ""),
                        "province": item.get("province", ""),
                        "city": item.get("city", ""),
                        "district": item.get("area", ""),
                        "distance": "",
                    })
            except Exception as e:
                return {"success": False, "error": f"请求失败: {str(e)[:50]}"}
    
    return {"success": True, "total": total, "data": results, "page": page, "page_size": page_size, "keyword": keyword, "location": location}




@app.post("/api/export-email")
async def export_email(data: dict, request: Request):
    """Send export file to user's email via configured SMTP."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return {"success": False, "error": "未登录"}
    import auth_db
    user = auth_db.validate_token(token)
    if not user:
        return {"success": False, "error": "登录已过期"}

    task_id = data.get("task_id", "")
    target_email = data.get("email", "").strip().lower()
    if not task_id or not target_email:
        return {"success": False, "error": "参数不完整"}
    if "@" not in target_email or "." not in target_email:
        return {"success": False, "error": "邮箱格式不正确"}

    task = tasks.get(task_id)
    if not task or task["status"] != "completed":
        return {"success": False, "error": "没有可导出的结果"}
    filepath = task.get("filepath")
    if not filepath or not os.path.exists(filepath):
        return {"success": False, "error": "文件不存在"}

    # Read SMTP config
    config_path = BASE_DIR / "data" / "email_config.json"
    if not config_path.exists():
        return {"success": False, "error": "邮件服务未配置，请联系管理员设置SMTP"}

    import json
    smtp_config = json.loads(config_path.read_text(encoding="utf-8"))
    host = smtp_config.get("host", "")
    port = int(smtp_config.get("port", 587))
    username = smtp_config.get("username", "")
    password = smtp_config.get("password", "")
    from_addr = smtp_config.get("from_addr", username)

    if not host or not username or not password:
        return {"success": False, "error": "SMTP 配置不完整"}

    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        from email.mime.base import MIMEBase
        from email import encoders

        msg = MIMEMultipart()
        msg["Subject"] = f"商家 POI 数据导出 - {task.get('filename', 'result.xlsx')}"
        msg["From"] = from_addr
        msg["To"] = target_email
        msg["Message-ID"] = f"<{uuid.uuid4().hex}@ameibar.com>"

        body_text = f"""您好，

您的商家 POI 数据导出已完成，请查收附件。

文件: {task.get('filename', 'result.xlsx')}
数据条数: {task.get('count', 0)}

—— 商家 POI 查询平台
"""
        msg.attach(MIMEText(body_text, "plain", "utf-8"))

        with open(filepath, "rb") as f:
            attachment = MIMEBase("application", "octet-stream")
            attachment.set_payload(f.read())
        encoders.encode_base64(attachment)
        attachment.add_header(
            "Content-Disposition",
            attachment.add_header("Content-Disposition", "attachment; filename=\\" + task.get("filename", "result.xlsx") + "\"")
        )
        msg.attach(attachment)

        _fname_for_attach = task.get("filename", "result.xlsx")
        with smtplib.SMTP(host, port, timeout=15) as server:
            server.starttls()
            server.login(username, password)
            server.send_message(msg)

        return {"success": True, "message": f"文件已发送至 {target_email}"}
    except smtplib.SMTPAuthenticationError:
        return {"success": False, "error": "SMTP 认证失败，请检查邮箱密码"}
    except smtplib.SMTPException as e:
        return {"success": False, "error": f"邮件发送失败: {str(e)[:80]}"}
    except Exception as e:
        return {"success": False, "error": f"发送失败: {str(e)[:80]}"}




import shutil

AVATAR_DIR = BASE_DIR / "data" / "avatars"
os.makedirs(str(AVATAR_DIR), exist_ok=True)

@app.post("/api/auth/avatar")
async def upload_avatar(request: Request):
    """Upload user avatar (base64 encoded image)."""
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return {"success": False, "error": "未登录"}
    import auth_db
    user = auth_db.validate_token(token)
    if not user:
        return {"success": False, "error": "登录已过期"}
    
    data = await request.json()
    image_data = data.get("image", "")
    if not image_data:
        return {"success": False, "error": "请选择头像图片"}
    
    # Decode base64
    import base64
    try:
        # Handle data:image/png;base64,xxx format
        if "," in image_data:
            image_data = image_data.split(",")[1]
        img_bytes = base64.b64decode(image_data)
    except Exception:
        return {"success": False, "error": "图片格式错误"}
    
    # Validate file size (max 2MB)
    if len(img_bytes) > 2 * 1024 * 1024:
        return {"success": False, "error": "图片不能超过2MB"}
    
    # Save file
    filepath = str(AVATAR_DIR / f"{user['id']}.jpg")
    with open(filepath, "wb") as f:
        f.write(img_bytes)
    
    return {"success": True, "message": "头像已更新", "avatar_url": f"/api/avatar/{user['id']}"}


@app.get("/api/avatar/{user_id}")
async def get_avatar(user_id: int):
    """Serve user avatar image."""
    for ext in ["jpg", "jpeg", "png", "gif", "webp"]:
        filepath = str(AVATAR_DIR / f"{user_id}.{ext}")
        if os.path.exists(filepath):
            media_type = f"image/{'jpeg' if ext in ('jpg','jpeg') else ext}"
            return FileResponse(filepath, media_type=media_type)
    return JSONResponse(status_code=404, content={"error": "Avatar not found"})


if __name__ == "__main__":
    import uvicorn
    print("=" * 50)
    print("  商家信息查询工具 — 高德地图 POI")
    print("=" * 50)
    print(f"  地址: http://localhost:{settings.PORT}")
    print(f"  高德 API: {'✅ 已配置' if get_amap_api_key() else '⚠ 未配置'}")
    print(f"  免费额度: 5000次/天")
    print(f"  批量搜索上限: {settings.MAX_RESULTS} 条")
    print("=" * 50)
    uvicorn.run(app, host=settings.HOST, port=settings.PORT)
